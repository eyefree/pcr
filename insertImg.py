import xlsxwriter
import os
from PIL import Image
from openpyxl import load_workbook

def insertImg():
    workbook = xlsxwriter.Workbook('image.xlsx')
    worksheet = workbook.add_worksheet()
    row = 1
    for root, dirs, files in os.walk(r"C:\Users\eyefree\Desktop\pcr\img"):
        files.sort(key=lambda x: int(x.split('.')[0]))
        for file in files:
            imgPath = os.path.join(root, file)
            #with Image.open(imgPath) as img:
                #width = img.width
                #height = img.height
            #缩小原图
            #width_30 = int(round(width * 0.3, 0))
            #img = Image.open(imgPath).convert('RGB')
            #wpercent = (width_30 / float(width))
            #hsize = int((float(height) * float(wpercent)))
            #img = img.resize((width_30, hsize), Image.ANTIALIAS)
            #img.save(imgPath)

            with Image.open(imgPath) as img:
                width = img.width
                height = img.height
            worksheet.set_column(row, 0, width / 7)
            worksheet.set_row(row, height / 4 * 3)
            worksheet.insert_image(row, 0, imgPath)
            print(imgPath)
            row += 1
    workbook.close()

#insertImg()
def resizeRow():
    workbook = load_workbook('pcr.xlsx')
    worksheet = workbook[workbook.sheetnames[0]]
    row = 1
    for root, dirs, files in os.walk(r"C:\Users\eyefree\Desktop\pcr\img"):
        files.sort(key=lambda x: int(x.split('.')[0]))
        for file in files:
            imgPath = os.path.join(root, file)
            with Image.open(imgPath) as img:
                h = img.height
            worksheet.row_dimensions[row].height = h
            row += 1
    workbook.save('pcr.xlsx')
resizeRow()
